#!/usr/bin/env python3
"""
Generate JSON snapshots of every tab in CA_INTER_G2_Planner_v1.xlsx.

These snapshots are bundled into the Astro site as a fallback:
- Used when SHEET_PUBLISHED_ID is empty (zero-config dev)
- Used when live CSV fetch fails (offline / 404)

Run:
    PYTHONPATH=/home/coder/.python_libs python3 scripts/generate-snapshots.py
"""
import json
import os
import re
import sys
from datetime import datetime, date

import openpyxl

HYPERLINK_RE = re.compile(r'=HYPERLINK\("([^"]+)"', re.IGNORECASE)

XLSX = "/home/coder/CA_INTER_G2_Planner_v1.xlsx"
OUT_DIR = os.path.join(os.path.dirname(__file__), "..", "src", "data", "snapshots")
HEADER_ROW = 5     # row index (1-based) where the actual column headers live
DATA_COL_START = 2 # column B; column A is the visual margin (always empty)


def cell_value(v):
    """Coerce openpyxl cell value to JSON-safe primitive."""
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return v
    s = str(v).strip()
    if not s:
        return None
    # Extract URL from =HYPERLINK("...","display") formulas
    m = HYPERLINK_RE.match(s)
    if m:
        return m.group(1)
    # Any other formula stays unevaluated in data_only=False; treat as null in
    # the snapshot so the UI renders "—" instead of the raw formula text.
    # Once the user publishes to Google Sheets, the CSV will have the computed value.
    if s.startswith("="):
        return None
    return s


def extract_tab(ws):
    """Return list[dict] starting from the first data row after HEADER_ROW."""
    # find header columns (anything non-empty in HEADER_ROW from DATA_COL_START)
    headers = []
    for c in range(DATA_COL_START, ws.max_column + 1):
        h = ws.cell(row=HEADER_ROW, column=c).value
        if h is not None and str(h).strip():
            headers.append((c, str(h).strip()))

    if not headers:
        return {"headers": [], "rows": []}

    rows = []
    for r in range(HEADER_ROW + 1, ws.max_row + 1):
        row_obj = {}
        any_value = False
        for col_idx, header in headers:
            v = cell_value(ws.cell(row=r, column=col_idx).value)
            if v is not None:
                any_value = True
            row_obj[header] = v
        if any_value:
            rows.append(row_obj)
    return {"headers": [h for _, h in headers], "rows": rows}


def extract_wellness(ws):
    """
    Wellness tab is laid out as section headers (column B) and bullet items
    (column C), starting from row 5. Output is { sections: [{title, items}, ...] }.
    """
    sections: list[dict] = []
    current: dict | None = None
    for r in range(HEADER_ROW, ws.max_row + 1):
        title = cell_value(ws.cell(row=r, column=2).value)
        item = cell_value(ws.cell(row=r, column=3).value)
        if title and not item:
            current = {"title": str(title), "items": []}
            sections.append(current)
        elif item and current is not None:
            current["items"].append(str(item))
    return {"sections": sections}


# Sheet name -> snapshot filename (kebab-case, no number prefix).
# Note: `11 · PDFs` snapshot is NOT generated here — it's owned by
# scripts/seed-pdfs-tab.py, which builds it from scripts/icai_inventory.py.
TAB_MAP = {
    "1 · Start Here": "start-here",
    "2 · Dashboard":  "dashboard",
    "3 · Syllabus":   "syllabus",
    "4 · Phases":     "phases",
    "5 · Daily Plan": "daily",
    "6 · Mocks Log":  "mocks",
    "7 · Key Dates":  "key-dates",
    "8 · Resources":  "resources",
    "9 · Wellness":   "wellness",
    "10 · Notes":     "notes",
}


def main():
    os.makedirs(OUT_DIR, exist_ok=True)
    # data_only=False so HYPERLINK formula text is visible for URL extraction
    wb = openpyxl.load_workbook(XLSX, data_only=False)

    summary = {}
    for sheet_name, slug in TAB_MAP.items():
        if sheet_name not in wb.sheetnames:
            print(f"WARN: missing tab {sheet_name}", file=sys.stderr)
            continue
        ws = wb[sheet_name]
        if slug == "wellness":
            data = extract_wellness(ws)
            count = sum(len(s["items"]) for s in data["sections"])
            summary[slug] = f'{len(data["sections"])} sections / {count} items'
        else:
            data = extract_tab(ws)
            summary[slug] = len(data["rows"])
        out_path = os.path.join(OUT_DIR, f"{slug}.json")
        with open(out_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        print(f"  {slug}: {summary[slug]} -> {out_path}")

    print()
    print("snapshot summary:", json.dumps(summary, indent=2))


if __name__ == "__main__":
    main()
